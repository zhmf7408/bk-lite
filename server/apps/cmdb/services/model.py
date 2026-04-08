import json
from dataclasses import asdict
from typing import Any

from apps.cmdb.constants.constants import (
    CLASSIFICATION,
    CREATE_MODEL_CHECK_ATTR,
    INST_NAME_INFOS,
    INSTANCE,
    MODEL,
    MODEL_ASSOCIATION,
    ORGANIZATION,
    SUBORDINATE_MODEL,
    UPDATE_MODEL_CHECK_ATTR_MAP,
    USER,
    OPERATOR_MODEL,
    DISPLAY_FIELD_CONFIG,
    ENUM_SELECT_MODE_DEFAULT,
)
from apps.cmdb.constants.field_constraints import TAG_ATTR_ID, TAG_MODE_FREE
from apps.cmdb.validators.field_validator import (
    normalize_tag_field_option as normalize_tag_field_option_config,
)
from apps.cmdb.validators import IdentifierValidator
from apps.cmdb.display_field.constants import DISPLAY_FIELD_TYPES, DISPLAY_SUFFIX
from apps.cmdb.graph.drivers.graph_client import GraphClient
from apps.cmdb.language.service import SettingLanguage
from apps.cmdb.models import UPDATE_INST, DELETE_INST, CREATE_INST, FieldGroup
from apps.cmdb.services.classification import ClassificationManage
from apps.cmdb.services.unique_rule import (
    UniqueRulePayload,
    apply_unique_rules_to_attr_export_rows,
    build_unique_rule_context,
    build_unique_rules_from_attr_rows,
    copy_unique_rules_to_model,
    create_unique_rule,
    delete_unique_rule,
    enrich_attrs_with_unique_display,
    guard_attr_change_against_unique_rules,
    list_unique_rule_candidate_fields,
    list_unique_rules,
    validate_unique_rules_against_existing_instances,
    update_unique_rule,
)
from apps.cmdb.utils.change_record import create_change_record
from apps.core.exceptions.base_app_exception import BaseAppException
from apps.core.services.user_group import UserGroup
from apps.rpc.system_mgmt import SystemMgmt
from apps.core.logger import cmdb_logger as logger

FIELD_GROUP_MANAGER: Any = getattr(FieldGroup, "objects")


class ModelManage(object):
    @staticmethod
    def _normalize_default_value(raw_value: Any) -> list[str]:
        if raw_value in (None, ""):
            return []

        source = raw_value if isinstance(raw_value, list) else [raw_value]
        seen: set[str] = set()
        normalized: list[str] = []
        for item in source:
            value = str(item or "").strip()
            if not value or value in seen:
                continue
            seen.add(value)
            normalized.append(value)
        return normalized

    @staticmethod
    def sanitize_attr_default_value(attr: dict, log_context: str = "") -> dict:
        item = dict(attr)
        normalized = ModelManage._normalize_default_value(item.get("default_value"))
        if item.get("attr_type") != "enum":
            item["default_value"] = normalized
            return item

        runtime_options = ModelManage.resolve_runtime_enum_options(item)
        valid_ids = {str(option.get("id")).strip() for option in runtime_options if isinstance(option, dict) and str(option.get("id", "")).strip()}
        sanitized = [value for value in normalized if value in valid_ids]
        select_mode = item.get("enum_select_mode", ENUM_SELECT_MODE_DEFAULT)
        if select_mode != "multiple":
            sanitized = sanitized[:1]

        removed_values = [value for value in normalized if value not in sanitized]
        if removed_values:
            logger.info(
                "[ModelDefaultValue] pruned stale defaults context=%s attr_id=%s removed=%s rule_type=%s",
                log_context or "runtime",
                item.get("attr_id"),
                removed_values,
                item.get("enum_rule_type", "custom"),
            )

        item["default_value"] = sanitized
        return item

    @staticmethod
    def _normalize_attr_constraints(attrs: list[dict]) -> list[dict]:
        normalized: list[dict] = []
        for attr in attrs:
            item = dict(attr)
            item.setdefault("is_only", False)
            item.setdefault("is_required", False)
            item.setdefault("editable", True)
            item.setdefault("option", {})
            item.setdefault("user_prompt", "")
            item.setdefault("default_value", [])
            item = ModelManage.sanitize_attr_default_value(item, log_context="normalize_attr_constraints")
            normalized.append(item)
        return normalized

    @staticmethod
    def _is_tag_attr(attr: dict) -> bool:
        return attr.get("attr_type") == "tag" or attr.get("attr_id") == TAG_ATTR_ID

    @staticmethod
    def validate_tag_attr_definition(attrs: list[dict], incoming_attr: dict) -> None:
        attr_type = incoming_attr.get("attr_type")
        incoming_attr_id = incoming_attr.get("attr_id")

        if attr_type == "tag" and incoming_attr_id != TAG_ATTR_ID:
            raise BaseAppException("tag 字段 attr_id 必须固定为 tag")

        if incoming_attr_id == TAG_ATTR_ID and attr_type != "tag":
            raise BaseAppException("attr_id 为 tag 的字段类型必须为 tag")

        if attr_type != "tag":
            return

        tag_count = sum(1 for attr in attrs if ModelManage._is_tag_attr(attr))
        if tag_count >= 1:
            raise BaseAppException("单模型最多允许一个 tag 字段")

    @staticmethod
    def normalize_tag_field_option(option: dict | list[Any] | None) -> dict:
        if isinstance(option, list):
            option = {"mode": TAG_MODE_FREE, "options": option}
        config = normalize_tag_field_option_config(option)
        return {
            "mode": config.mode,
            "options": [{"key": item.key, "value": item.value} for item in config.options],
        }

    @staticmethod
    def merge_tag_options_from_values(model_id: str, values: list[str]) -> None:
        if not values:
            return
        model_info = ModelManage.search_model_info(model_id)
        if not model_info:
            return
        attrs = ModelManage.parse_attrs(model_info.get("attrs", "[]"))
        tag_attr = next((attr for attr in attrs if ModelManage._is_tag_attr(attr)), None)
        if not tag_attr:
            return

        option = ModelManage.normalize_tag_field_option(tag_attr.get("option") or {})
        if option.get("mode") != TAG_MODE_FREE:
            return

        existing = {
            (str(item.get("key", "")).strip(), str(item.get("value", "")).strip()) for item in option.get("options", []) if isinstance(item, dict)
        }
        changed = False
        for raw in values:
            if not isinstance(raw, str) or raw.count(":") != 1:
                continue
            key, value = [part.strip() for part in raw.split(":", 1)]
            if not key or not value:
                continue
            pair = (key, value)
            if pair in existing:
                continue
            existing.add(pair)
            option["options"].append({"key": key, "value": value})
            changed = True

        if not changed:
            return

        for attr in attrs:
            if ModelManage._is_tag_attr(attr):
                attr["option"] = option
                break

        with GraphClient() as ag:
            ag.set_entity_properties(
                MODEL,
                [model_info["_id"]],
                {"attrs": json.dumps(attrs, ensure_ascii=False)},
                {},
                [],
                False,
            )

    @staticmethod
    def _validate_attr_id(attr_id: str):
        if not IdentifierValidator.is_valid(attr_id):
            raise BaseAppException(IdentifierValidator.get_error_message("属性ID"))

    @staticmethod
    def _validate_model_id(model_id: str):
        if not IdentifierValidator.is_valid(model_id):
            raise BaseAppException(IdentifierValidator.get_error_message("模型ID"))

    @staticmethod
    def normalize_enum_public_binding(attr_info: dict, current_attr: dict | None = None) -> dict:
        """
        规范化 enum 公共选项库绑定信息并回填 option 快照。

        处理逻辑：
        1. 若 attr_type != enum，直接返回原 attr_info，不做处理。
        2. enum_rule_type 默认为 custom。
        3. 若 enum_rule_type == public_library：
           - 校验 public_library_id 必填
           - 从公共库拉取最新 options 并写入 option 字段（快照）
        4. 若 enum_rule_type == custom：
           - 清空 public_library_id
           - 保持 option 不变

        Args:
            attr_info: 前端传入的属性配置
            current_attr: 当前已存在的属性（更新场景下传入）

        Returns:
            dict: 规范化后的 attr_info（原地修改并返回）

        Raises:
            BaseAppException: 公共库不存在或配置不合法时抛出
        """
        if attr_info.get("attr_type") != "enum":
            return attr_info

        option_value = attr_info.get("option")
        if isinstance(option_value, dict) and option_value.get("enum_rule_type"):
            attr_info["enum_rule_type"] = option_value.get("enum_rule_type", "custom")
            attr_info["public_library_id"] = option_value.get("public_library_id")
            if "enum_select_mode" in option_value:
                attr_info["enum_select_mode"] = option_value.get("enum_select_mode")
            attr_info["option"] = option_value.get("option", [])

        enum_rule_type = attr_info.get("enum_rule_type", "custom")
        attr_info["enum_rule_type"] = enum_rule_type

        if enum_rule_type == "public_library":
            public_library_id = attr_info.get("public_library_id")
            if not public_library_id:
                raise BaseAppException("绑定公共选项库时 public_library_id 必填")

            from apps.cmdb.services.public_enum_library import get_library_or_raise

            library = get_library_or_raise(public_library_id)
            attr_info["option"] = library.options
            attr_info["public_library_id"] = public_library_id

            logger.info(
                f"[EnumPublicBinding] normalized attr_id={attr_info.get('attr_id')}, "
                f"enum_rule_type={enum_rule_type}, public_library_id={public_library_id}"
            )
        else:
            attr_info["enum_rule_type"] = "custom"
            attr_info["public_library_id"] = None

        return attr_info

    @staticmethod
    def validate_enum_rule_immutable(current_attr: dict, incoming_attr: dict) -> None:
        """
        校验字段创建后 enum_rule_type 不可切换。

        规则：
        - 仅对 attr_type == enum 的字段生效
        - 创建时：任意 enum_rule_type（custom / public_library）均可
        - 更新时：不允许从 custom 切换到 public_library，或反之

        Args:
            current_attr: 当前已存储的属性定义
            incoming_attr: 本次请求传入的属性定义

        Raises:
            BaseAppException: 规则类型切换时抛出
        """
        if current_attr.get("attr_type") != "enum":
            return
        if incoming_attr.get("attr_type") != "enum":
            return

        current_rule = current_attr.get("enum_rule_type", "custom")
        incoming_rule = incoming_attr.get("enum_rule_type", "custom")

        if current_rule != incoming_rule:
            raise BaseAppException(f"枚举字段创建后规则类型不可切换（当前: {current_rule}）")

    @staticmethod
    def ensure_enum_select_mode(attr_info: dict) -> dict:
        """
        确保枚举字段具有 enum_select_mode 属性。

        规则：
        - 仅对 attr_type == enum 的字段生效
        - 若未提供 enum_select_mode，则默认设置为 single

        Args:
            attr_info: 属性配置字典

        Returns:
            dict: 规范化后的 attr_info（原地修改并返回）
        """
        if attr_info.get("attr_type") != "enum":
            return attr_info

        if "enum_select_mode" not in attr_info:
            attr_info["enum_select_mode"] = ENUM_SELECT_MODE_DEFAULT

        return attr_info

    @staticmethod
    def validate_enum_select_mode_immutable(current_attr: dict, incoming_attr: dict) -> None:
        """
        校验字段创建后 enum_select_mode 不可切换。

        规则：
        - 仅对 attr_type == enum 的字段生效
        - 创建时：任意 enum_select_mode（single / multiple）均可
        - 更新时：不允许从 single 切换到 multiple，或反之

        Args:
            current_attr: 当前已存储的属性定义
            incoming_attr: 本次请求传入的属性定义

        Raises:
            BaseAppException: 选择模式切换时抛出
        """
        if current_attr.get("attr_type") != "enum":
            return
        if incoming_attr.get("attr_type") != "enum":
            return

        current_mode = current_attr.get("enum_select_mode", ENUM_SELECT_MODE_DEFAULT)
        incoming_mode = incoming_attr.get("enum_select_mode", current_mode)

        if current_mode != incoming_mode:
            raise BaseAppException(f"枚举字段创建后选择模式不可切换（当前: {current_mode}）")

    @staticmethod
    def resolve_runtime_enum_options(attr: dict) -> list[dict]:
        """
        运行时解析枚举选项。

        口径：
        - 若 enum_rule_type == public_library，优先从公共库实时拉取
        - 若公共库不存在或查询失败，回退到字段快照（attr.option）
        - 若 enum_rule_type == custom，直接返回 attr.option

        Args:
            attr: 字段属性定义

        Returns:
            list[dict]: 枚举选项列表 [{"id": str, "name": str}, ...]
        """
        if attr.get("attr_type") != "enum":
            return []

        enum_rule_type = attr.get("enum_rule_type", "custom")
        option = attr.get("option", [])

        if enum_rule_type != "public_library":
            return option if isinstance(option, list) else []

        public_library_id = str(attr.get("public_library_id") or "").strip()
        if not public_library_id:
            return option if isinstance(option, list) else []

        try:
            from apps.cmdb.services.public_enum_library import get_library_or_raise

            library = get_library_or_raise(public_library_id)
            runtime_options = library.options
            return runtime_options if isinstance(runtime_options, list) else []
        except Exception as e:
            logger.warning(f"[EnumPublicBinding] resolve_runtime_enum_options fallback to snapshot, public_library_id={public_library_id}, error={e}")
            return option if isinstance(option, list) else []

    @staticmethod
    def _add_display_field_to_attrs(attrs: list, attr_info: dict, model_id: str, is_pre: bool = False):
        """
        为指定属性添加 _display 字段定义
        Args:
            attrs: 属性列表
            attr_info: 属性信息字典
            model_id: 模型ID
            is_pre: 是否为预定义字段
        Returns:
            bool: 是否添加了 _display 字段
        """
        from apps.cmdb.display_field import DISPLAY_FIELD_TYPES, DISPLAY_SUFFIX

        attr_type = attr_info.get("attr_type")
        if attr_type not in DISPLAY_FIELD_TYPES:
            return False

        attr_id = attr_info["attr_id"]
        display_field_id = f"{attr_id}{DISPLAY_SUFFIX}"

        # 检查是否已存在 _display 字段
        if display_field_id in {a["attr_id"] for a in attrs}:
            return False

        # 创建 _display 字段定义
        display_field = {
            "attr_id": display_field_id,
            "attr_name": attr_info.get("attr_name"),
            "attr_group": attr_info.get("attr_group", "default"),
            "group_id": attr_info.get("group_id"),
            "model_id": model_id,
            **DISPLAY_FIELD_CONFIG,
        }
        if is_pre:
            display_field["is_pre"] = True

        attrs.append(display_field)
        return True

    @staticmethod
    def _remove_display_field_from_attrs(attrs: list, attr_id: str):
        """
        从属性列表中删除指定属性的 _display 字段定义
        Args:
            attrs: 属性列表
            attr_id: 原始属性ID
        Returns:
            tuple: (新的属性列表, 是否删除了 _display 字段)
        """
        from apps.cmdb.display_field import DisplayFieldHandler

        display_field_id = f"{attr_id}{DISPLAY_SUFFIX}"
        original_count = len(attrs)
        new_attrs = [a for a in attrs if a["attr_id"] != display_field_id]

        return new_attrs, len(new_attrs) < original_count

    @staticmethod
    def _handle_attr_type_change(
        attrs: list,
        attr_id: str,
        old_type: str,
        new_type: str,
        attr_info: dict,
        model_id: str,
        graph_client,
    ):
        """
        处理属性类型变更时的 _display 字段逻辑
        Args:
            attrs: 属性列表
            attr_id: 属性ID
            old_type: 原类型
            new_type: 新类型
            attr_info: 属性信息
            model_id: 模型ID
            graph_client: 图数据库客户端实例
        Returns:
            list: 更新后的属性列表
        """
        from apps.cmdb.display_field import DisplayFieldHandler

        display_field_id = f"{attr_id}{DISPLAY_SUFFIX}"

        # 情况1: 从非目标类型改为目标类型 -> 添加 _display 字段
        if old_type not in DISPLAY_FIELD_TYPES and new_type in DISPLAY_FIELD_TYPES:
            ModelManage._add_display_field_to_attrs(attrs, attr_info, model_id)

        # 情况2: 从目标类型改为非目标类型 -> 删除 _display 字段和实例数据
        elif old_type in DISPLAY_FIELD_TYPES and new_type not in DISPLAY_FIELD_TYPES:
            # 从 attrs 中删除 _display 字段定义
            attrs, removed = ModelManage._remove_display_field_from_attrs(attrs, attr_id)

            # 删除所有实例的 _display 字段数据
            if removed:
                model_params = [{"field": "model_id", "type": "str=", "value": model_id}]
                graph_client.remove_entitys_properties(INSTANCE, model_params, [display_field_id])

        return attrs

    @staticmethod
    def create_model(data: dict, username="admin"):
        """
        创建模型
        """

        attrs = list(INST_NAME_INFOS)  # 复制默认字段列表
        model_id = str(data.get("model_id", ""))

        # 为默认字段中的目标类型添加 _display 字段定义
        for attr in list(attrs):  # 使用 list() 避免迭代时修改列表
            ModelManage._add_display_field_to_attrs(attrs, attr, model_id, is_pre=True)

        data.update(attrs=json.dumps(attrs), unique_rules="[]")

        with GraphClient() as ag:
            exist_items, _ = ag.query_entity(MODEL, [])
            result = ag.create_entity(MODEL, data, CREATE_MODEL_CHECK_ATTR, exist_items)
            classification_info = ClassificationManage.search_model_classification_info(data["classification_id"])
            _ = ag.create_edge(
                SUBORDINATE_MODEL,
                classification_info["_id"],
                CLASSIFICATION,
                result["_id"],
                MODEL,
                dict(
                    classification_model_asst_id=f"{result['classification_id']}_{SUBORDINATE_MODEL}_{result['model_id']}"
                    # noqa
                ),
                "classification_model_asst_id",
            )

        # 初始化排除字段缓存
        from apps.cmdb.display_field import ExcludeFieldsCache

        ExcludeFieldsCache.update_on_model_change(data["model_id"])

        # 为新建模型创建默认字段分组
        FIELD_GROUP_MANAGER.create(
            model_id=data["model_id"],
            group_name="default",
            order=1,
            is_collapsed=False,
            description="默认分组",
            created_by=username,
            attr_orders=[attr.get("attr_id") for attr in attrs if not attr.get("is_display_field")],
        )

        create_change_record(
            operator=username,
            model_id=data["model_id"],
            label="模型管理",
            _type=CREATE_INST,
            message=f"创建模型. 模型名称: {data['model_name']}",
            inst_id=result["_id"],
            model_object=OPERATOR_MODEL,
        )
        return result

    @staticmethod
    def copy_model(
        src_model_id: str,
        new_model_id: str,
        new_model_name: str,
        classification_id: str | None = None,
        group: list | None = None,
        icn: str | None = None,
        copy_attributes: bool = False,
        copy_relationships: bool = False,
        username="admin",
    ):
        """
        复制模型
        Args:
            src_model_id: 源模型ID
            new_model_id: 新模型ID
            new_model_name: 新模型名称
            classification_id: 模型分类ID（可选，不传则继承源模型）
            group: 组织列表（可选，不传则继承源模型）
            icn: 图标（可选，不传则继承源模型）
            copy_attributes: 是否复制属性
            copy_relationships: 是否复制关系
            username: 操作用户
        Returns:
            新模型信息
        """
        # 校验新模型ID格式
        ModelManage._validate_model_id(new_model_id)

        # 校验复制方式
        if not copy_attributes and not copy_relationships:
            raise BaseAppException("至少选择一种复制方式（属性或关系）")

        # 获取源模型信息
        src_model_info = ModelManage.search_model_info(src_model_id)
        if not src_model_info:
            raise BaseAppException("源模型不存在")

        # 准备属性列表
        if copy_attributes:
            # 完全复制源模型的所有属性
            src_attrs = ModelManage.parse_attrs(src_model_info.get("attrs", "[]"))
            attrs = [dict(attr) for attr in src_attrs]
        else:
            # 不复制属性：只使用默认属性
            attrs = list(INST_NAME_INFOS)
            # 为默认字段中的目标类型添加 _display 字段定义
            for attr in list(attrs):
                ModelManage._add_display_field_to_attrs(attrs, attr, new_model_id, is_pre=True)

        # 构建新模型数据
        new_model_data = {
            "model_id": new_model_id,
            "model_name": new_model_name,
            "classification_id": classification_id or src_model_info.get("classification_id"),
            "group": group if group is not None else src_model_info.get("group", []),
            "icn": icn if icn is not None else src_model_info.get("icn", ""),
            "attrs": json.dumps(attrs),
            "unique_rules": "[]",
        }

        # 一次性创建模型（包含所有属性）
        with GraphClient() as ag:
            exist_items, _ = ag.query_entity(MODEL, [])
            new_model = ag.create_entity(MODEL, new_model_data, CREATE_MODEL_CHECK_ATTR, exist_items)

            # 创建模型与分类的关联
            classification_info = ClassificationManage.search_model_classification_info(new_model_data["classification_id"])
            ag.create_edge(
                SUBORDINATE_MODEL,
                classification_info["_id"],
                CLASSIFICATION,
                new_model["_id"],
                MODEL,
                dict(classification_model_asst_id=f"{new_model['classification_id']}_{SUBORDINATE_MODEL}_{new_model['model_id']}"),
                "classification_model_asst_id",
            )

        # 初始化排除字段缓存
        from apps.cmdb.display_field import ExcludeFieldsCache

        ExcludeFieldsCache.update_on_model_change(new_model_id)

        try:
            # 处理字段分组复制
            if copy_attributes:
                # 复制源模型的字段分组配置
                src_field_groups = FIELD_GROUP_MANAGER.filter(model_id=src_model_id).order_by("order")
                for src_group in src_field_groups:
                    FIELD_GROUP_MANAGER.create(
                        model_id=new_model_id,
                        group_name=src_group.group_name,
                        attr_orders=src_group.attr_orders or [],
                        order=src_group.order,
                    )
                copy_unique_rules_to_model(src_model_id, new_model_id, username)
            else:
                # 不复制属性时，为新模型创建默认分组
                FIELD_GROUP_MANAGER.create(
                    model_id=new_model_id,
                    group_name="default",
                    order=1,
                    is_collapsed=False,
                    description="默认分组",
                    created_by=username,
                    attr_orders=[attr.get("attr_id") for attr in attrs if not attr.get("is_display_field")],
                )

            # 处理关系复制
            if copy_relationships:
                associations = ModelManage.model_association_search(src_model_id)

                for assoc in associations:
                    # 确定新的源模型ID和目标模型ID
                    new_src_model_id = new_model_id if assoc["src_model_id"] == src_model_id else assoc["src_model_id"]
                    new_dst_model_id = new_model_id if assoc["dst_model_id"] == src_model_id else assoc["dst_model_id"]

                    # 如果关联的两端都是源模型（自关联），则两端都改为新模型
                    if assoc["src_model_id"] == src_model_id and assoc["dst_model_id"] == src_model_id:
                        new_src_model_id = new_model_id
                        new_dst_model_id = new_model_id

                    # 获取新的src和dst的_id
                    src_model_info_new = ModelManage.search_model_info(new_src_model_id)
                    dst_model_info_new = ModelManage.search_model_info(new_dst_model_id)

                    if not src_model_info_new or not dst_model_info_new:
                        continue

                    # 创建新的关联ID
                    new_model_asst_id = f"{new_src_model_id}_{assoc['asst_id']}_{new_dst_model_id}"

                    # 复制关联关系
                    try:
                        ModelManage.model_association_create(
                            src_id=src_model_info_new["_id"],
                            dst_id=dst_model_info_new["_id"],
                            src_model_id=new_src_model_id,
                            dst_model_id=new_dst_model_id,
                            asst_id=assoc.get("asst_id", ""),
                            asst_name=assoc.get("asst_name", ""),
                            mapping=assoc.get("mapping", "1:n"),
                            on_delete=assoc.get("on_delete", "none"),
                            is_pre=assoc.get("is_pre", False),
                            model_asst_id=new_model_asst_id,
                        )
                    except BaseAppException:
                        # 如果关联已存在，跳过
                        continue

            # 记录变更
            create_change_record(
                operator=username,
                model_id=new_model_id,
                label="模型管理",
                _type=CREATE_INST,
                message=f"复制模型. 源模型: {src_model_info['model_name']}, 新模型: {new_model_name}",
                inst_id=new_model["_id"],
                model_object=OPERATOR_MODEL,
            )

            return new_model

        except Exception as e:
            # 如果复制过程中出错，删除已创建的模型
            try:
                ModelManage.delete_model(new_model["_id"])
            except Exception:  # noqa: BLE001 - 清理失败不应掩盖原始错误
                pass
            raise e

    @staticmethod
    def delete_model(id: int):
        """
        删除模型
        """
        with GraphClient() as ag:
            ag.batch_delete_entity(MODEL, [id])

    @staticmethod
    def update_model(id: int, data: dict):
        """
        更新模型
        TODO 不能单独更新一个字段，如只更新icon，传递全部字段会导致其他字段校验不通过 model_name 后续考虑优化
        """
        model_id = data.pop("model_id", "")  # 不能更新model_id
        with GraphClient() as ag:
            exist_items, _ = ag.query_entity(MODEL, [{"field": "model_id", "type": "str<>", "value": model_id}])
            # 排除当前正在更新的模型，避免自己和自己比较
            exist_items = [i for i in exist_items if i["_id"] != id]
            model = ag.set_entity_properties(MODEL, [id], data, UPDATE_MODEL_CHECK_ATTR_MAP, exist_items)
        return model[0]

    @staticmethod
    def search_model(
        language: str = "en",
        order_type: str = "ASC",
        order: str = "id",
        permissions_map: dict | None = None,
        classification_ids: list | None = None,
        creator: str = "",
    ):
        """
        查询模型
        Args:
            language: 语言，默认英语
            order_type: 排序方式，asc升序/desc降序
            order: 排序字段，默认order_id
            classification_ids: 分类ID列表，可选，用于过滤特定分类下的模型
            permissions_map: 权限过滤字典
            creator: 创建人，可选，用于过滤特定创建人的模型
        """

        permissions_map = permissions_map or {}
        format_permission_dict = {}

        for organization_id, organization_permission_data in permissions_map.items():
            _query_list = []
            if classification_ids:
                _query_list.append(
                    {
                        "field": "classification_id",
                        "type": "str[]",
                        "value": classification_ids,
                    }
                )
            model_ids = organization_permission_data["inst_names"]
            if model_ids:
                _query_list.append({"field": "model_id", "type": "str[]", "value": model_ids})
                if creator:
                    # 只有创建人条件
                    _query_list.append({"field": "_creator", "type": "str=", "value": creator})

            format_permission_dict[organization_id] = _query_list

        with GraphClient() as ag:
            query = dict(
                label=MODEL,
                params=[],
                order=order,
                order_type=order_type,
                format_permission_dict=format_permission_dict,
                param_type="OR",
                organization_field="group",
            )
            models, _ = ag.query_entity(**query)

        lan = SettingLanguage(language)

        for model in models:
            model["model_name"] = lan.get_val("MODEL", model["model_id"]) or model["model_name"]
            # 确保所有模型都有order_id
            if "order_id" not in model:
                model["order_id"] = 0

        return models

    @staticmethod
    def parse_attrs(attrs: str):
        return json.loads(attrs.replace('\\"', '"'))

    @staticmethod
    def create_model_attr(model_id, attr_info, username="admin"):
        """
        创建模型属性
        """
        with GraphClient() as ag:
            if attr_info.get("attr_type") == "tag":
                attr_info["attr_id"] = TAG_ATTR_ID
                attr_info["editable"] = True
                attr_info["is_only"] = False
                attr_info["is_required"] = False
                attr_info["option"] = ModelManage.normalize_tag_field_option(attr_info.get("option") or {})

            if attr_info.get("attr_type") == "enum":
                ModelManage.normalize_enum_public_binding(attr_info)
                ModelManage.ensure_enum_select_mode(attr_info)

            attr_info = ModelManage.sanitize_attr_default_value(attr_info, log_context="create_model_attr")

            ModelManage._validate_attr_id(attr_info["attr_id"])
            model_query = {"field": "model_id", "type": "str=", "value": model_id}
            models, _ = ag.query_entity(MODEL, [model_query])
            model_count = len(models)
            if model_count == 0:
                raise BaseAppException("model not present")
            model_info = models[0]
            attrs = ModelManage.parse_attrs(model_info.get("attrs", "[]"))
            ModelManage.validate_tag_attr_definition(attrs, attr_info)
            if attr_info["attr_id"] in {i["attr_id"] for i in attrs}:
                raise BaseAppException("model attr repetition")
            attrs.append(attr_info)

            ModelManage._add_display_field_to_attrs(attrs, attr_info, model_id)

            result = ag.set_entity_properties(MODEL, [model_info["_id"]], dict(attrs=json.dumps(attrs)), {}, [], False)

        # 更新排除字段缓存
        from apps.cmdb.display_field import ExcludeFieldsCache

        updated_attrs = ModelManage.parse_attrs(result[0].get("attrs", "[]"))
        ExcludeFieldsCache.update_on_model_change(model_id)

        attrs = updated_attrs

        attr = None
        for attr in attrs:
            if attr["attr_id"] != attr_info["attr_id"]:
                continue
            attr = attr

        create_change_record(
            operator=username,
            model_id=model_id,
            label="模型管理",
            _type=CREATE_INST,
            message=f"创建模型属性. 模型名称: {model_info['model_name']}",
            inst_id=model_info["_id"],
            model_object=OPERATOR_MODEL,
        )

        return attr

    @staticmethod
    def update_model_attr(model_id, attr_info, username="admin"):
        """
        更新模型属性
        """
        with GraphClient() as ag:
            ModelManage._validate_attr_id(attr_info["attr_id"])
            model_query = {"field": "model_id", "type": "str=", "value": model_id}
            models, model_count = ag.query_entity(MODEL, [model_query])
            if model_count == 0:
                raise BaseAppException("model not present")
            model_info = models[0]
            attrs = ModelManage.parse_attrs(model_info.get("attrs", "[]"))
            if attr_info["attr_id"] not in {i["attr_id"] for i in attrs}:
                raise BaseAppException("model attr not present")

            current_attr = next(
                (attr for attr in attrs if attr["attr_id"] == attr_info["attr_id"]),
                None,
            )
            if not current_attr:
                raise BaseAppException("model attr not present")

            if current_attr.get("is_required") != attr_info.get("is_required"):
                guard_attr_change_against_unique_rules(
                    model_id,
                    attr_info["attr_id"],
                    attr_info,
                    "update_required",
                    username,
                )
            if current_attr.get("attr_type") != attr_info.get("attr_type"):
                guard_attr_change_against_unique_rules(
                    model_id,
                    attr_info["attr_id"],
                    attr_info,
                    "update_type",
                    username,
                )

            is_tag_attr = ModelManage._is_tag_attr(current_attr)
            if is_tag_attr:
                if attr_info.get("attr_id") != TAG_ATTR_ID:
                    raise BaseAppException("tag 字段 attr_id 不允许修改")
                attr_info["attr_type"] = "tag"
                attr_info["editable"] = True
                attr_info["is_only"] = False
                attr_info["is_required"] = False
                attr_info["option"] = ModelManage.normalize_tag_field_option(attr_info.get("option") or current_attr.get("option") or {})

            is_enum_attr = current_attr.get("attr_type") == "enum"
            if is_enum_attr:
                ModelManage.validate_enum_rule_immutable(current_attr, attr_info)
                ModelManage.validate_enum_select_mode_immutable(current_attr, attr_info)
                ModelManage.normalize_enum_public_binding(attr_info, current_attr)

            attr_info = ModelManage.sanitize_attr_default_value(attr_info, log_context="update_model_attr")

            for attr in attrs:
                if attr_info["attr_id"] != attr["attr_id"]:
                    continue
                attr.update(
                    attr_group=attr_info["attr_group"],
                    attr_name=attr_info["attr_name"],
                    is_required=False if is_tag_attr else attr_info["is_required"],
                    editable=True if is_tag_attr else attr_info["editable"],
                    option=attr_info["option"],
                    user_prompt=attr_info["user_prompt"],
                    default_value=attr_info.get("default_value", []),
                )
                if is_enum_attr:
                    attr["enum_rule_type"] = attr_info.get("enum_rule_type", "custom")
                    attr["public_library_id"] = attr_info.get("public_library_id")
                    attr["enum_select_mode"] = current_attr.get("enum_select_mode", ENUM_SELECT_MODE_DEFAULT)

            result = ag.set_entity_properties(MODEL, [model_info["_id"]], dict(attrs=json.dumps(attrs)), {}, [], False)

        attrs = ModelManage.parse_attrs(result[0].get("attrs", "[]"))

        # 更新排除字段缓存
        from apps.cmdb.display_field import ExcludeFieldsCache

        ExcludeFieldsCache.update_on_model_change(model_id)

        attr = None
        for attr in attrs:
            if attr["attr_id"] == attr_info["attr_id"]:
                attr = attr
                break

        create_change_record(
            operator=username,
            model_id=model_id,
            label="模型管理",
            _type=UPDATE_INST,
            message=f"修改模型属性. 模型名称: {model_info['model_name']}",
            inst_id=model_info["_id"],
            model_object=OPERATOR_MODEL,
        )

        return attr

    @staticmethod
    def update_enum_instances_display(model_id: str, attr_id: str, new_options: list):
        """
        更新枚举类型字段变更后所有实例的 _display 冗余字段
        Args:
            model_id: 模型ID
            attr_id: 属性ID
            new_options: 新的枚举选项列表
        Returns:
            int: 更新的实例数量
        """
        from apps.cmdb.display_field import DisplayFieldConverter

        updated_count = 0
        display_field_id = f"{attr_id}_display"

        try:
            with GraphClient() as ag:
                # 查询该模型的所有实例
                instances, _ = ag.query_entity(INSTANCE, [{"field": "model_id", "type": "str=", "value": model_id}])

                # 批量更新实例的 _display 字段
                for instance in instances:
                    # 检查实例是否有该枚举字段
                    if attr_id in instance and instance[attr_id]:
                        enum_value = instance[attr_id]

                        # 使用统一的转换器生成新的 _display 值
                        new_display_value = DisplayFieldConverter.convert_enum(enum_value, new_options)

                        # 更新实例的 _display 字段
                        update_data = {display_field_id: new_display_value}
                        ag.batch_update_node_properties(INSTANCE, [instance["_id"]], update_data)
                        updated_count += 1

                if updated_count > 0:
                    logger.info(
                        f"[update_enum_instances_display] 枚举选项变更，已更新 {updated_count} 个实例的 {display_field_id} 字段, "
                        f"模型: {model_id}, 字段: {attr_id}"
                    )

        except Exception as e:
            logger.error(
                f"[update_enum_instances_display] 更新实例枚举 _display 字段失败: 模型={model_id}, 字段={attr_id}, 错误={e}",
                exc_info=True,
            )
            # 不抛出异常，避免中断主流程

        return updated_count

    @staticmethod
    def delete_model_attr(model_id: str, attr_id: str, username: str = "admin"):
        """
        删除模型属性
        """
        guard_attr_change_against_unique_rules(
            model_id,
            attr_id,
            None,
            "delete",
            username,
        )
        with GraphClient() as ag:
            model_query = {"field": "model_id", "type": "str=", "value": model_id}
            models, model_count = ag.query_entity(MODEL, [model_query])
            if model_count == 0:
                raise BaseAppException("model not present")
            model_info = models[0]
            attrs = ModelManage.parse_attrs(model_info.get("attrs", "[]"))

            # 检查要删除的字段类型,如果是目标类型,也需要删除对应的 _display 字段
            from apps.cmdb.display_field import DisplayFieldHandler

            fields_to_remove = [attr_id]

            # 检查是否需要同时删除 _display 字段
            for attr in attrs:
                if attr["attr_id"] == attr_id and attr.get("attr_type") in DISPLAY_FIELD_TYPES:
                    display_field_id = f"{attr_id}{DISPLAY_SUFFIX}"
                    fields_to_remove.append(display_field_id)
                    break

            new_attrs = [attr for attr in attrs if attr["attr_id"] not in fields_to_remove]
            result = ag.set_entity_properties(
                MODEL,
                [model_info["_id"]],
                dict(attrs=json.dumps(new_attrs)),
                {},
                [],
                False,
            )

            # 模型属性删除后，要删除对应模型实例的属性(包括 _display 字段)
            model_params = [{"field": "model_id", "type": "str=", "value": model_id}]
            ag.remove_entitys_properties(INSTANCE, model_params, fields_to_remove)

        # 更新排除字段缓存
        from apps.cmdb.display_field import ExcludeFieldsCache

        updated_attrs = ModelManage.parse_attrs(result[0].get("attrs", "[]"))
        ExcludeFieldsCache.update_on_model_change(model_id)

        create_change_record(
            operator=username,
            model_id=model_id,
            label="模型管理",
            _type=DELETE_INST,
            message=f"删除模型属性. 模型名称: {model_info['model_name']}",
            inst_id=model_info["_id"],
            model_object=OPERATOR_MODEL,
        )

        return updated_attrs

    @staticmethod
    def search_model_info(model_id: str):
        """
        查询模型详情
        """
        query_data = [{"field": "model_id", "type": "str=", "value": model_id}]
        with GraphClient() as ag:
            models, _ = ag.query_entity(MODEL, query_data)
        if len(models) == 0:
            return {}

        model = models[0]

        # if not display_field:
        #     return model
        #
        # # 过滤掉 is_display_field 为 true 的字段
        # if "attrs" in model and model["attrs"]:
        #     attrs = ModelManage.parse_attrs(model["attrs"])
        #     filtered_attrs = [attr for attr in attrs if not attr.get("is_display_field")]
        #     model["attrs"] = json.dumps(filtered_attrs, ensure_ascii=False)

        return model

    @staticmethod
    def get_organization_option(items: list, result: list, name_prefix: str = ""):
        for item in items:
            if name_prefix:
                name = f"{name_prefix}/{item['name']}"
            else:
                name = item["name"]
            result.append(
                dict(
                    id=item["id"],
                    name=name,
                    is_default=False,
                    type="str",
                )
            )
            if item["subGroups"]:
                ModelManage.get_organization_option(item["subGroups"], result, name)

    @staticmethod
    def search_model_attr(model_id: str, language: str = "en"):
        """
        查询模型属性
        """
        model_info = ModelManage.search_model_info(model_id)
        attrs = ModelManage._normalize_attr_constraints(ModelManage.parse_attrs(model_info.get("attrs", "[]")))
        attrs = [ModelManage.sanitize_attr_default_value(attr, log_context="search_model_attr") for attr in attrs]
        unique_rules = build_unique_rule_context(model_id).unique_rules
        # TODO 语言包
        # lan = SettingLanguage(language)
        # model_attr = lan.get_val("ATTR", model_id)
        # for attr in attrs:
        #     if model_attr:
        #         attr["attr_name"] = model_attr.get(attr["attr_id"]) or attr["attr_name"]
        #
        return enrich_attrs_with_unique_display(attrs, unique_rules, model_id)

    @staticmethod
    def search_model_attr_v2(model_id: str):
        """
        查询模型属性
        """
        model_info = ModelManage.search_model_info(model_id)
        attrs = ModelManage._normalize_attr_constraints(ModelManage.parse_attrs(model_info.get("attrs", "[]")))
        attrs = [ModelManage.sanitize_attr_default_value(attr, log_context="search_model_attr_v2") for attr in attrs]
        unique_rules = build_unique_rule_context(model_id).unique_rules
        attr_types = {attr["attr_type"] for attr in attrs}
        system_mgmt_client = SystemMgmt()

        if ORGANIZATION in attr_types:
            groups = UserGroup.get_all_groups(system_mgmt_client)
            # 获取默认的第一个根组织
            groups = groups if groups else []
            option = []
            ModelManage.get_organization_option(groups, option)
            for attr in attrs:
                if attr["attr_type"] == ORGANIZATION:
                    attr.update(option=option)

        if USER in attr_types:
            users = UserGroup.get_all_users(system_mgmt_client)
            option = [
                dict(
                    # id=user["username"],
                    id=user["id"],
                    name=user["username"],
                    username=user.get("username"),
                    display_name=user.get("display_name"),
                    is_default=False,
                    type="str",
                )
                for user in users["users"]
            ]
            for attr in attrs:
                if attr["attr_type"] == USER:
                    attr.update(option=option)

        return enrich_attrs_with_unique_display(attrs, unique_rules, model_id)

    @staticmethod
    def get_model_unique_rules(model_id: str, editing_rule_id: str | None = None):
        return {
            "rules": list_unique_rules(model_id),
            "candidate_fields": [asdict(candidate) for candidate in list_unique_rule_candidate_fields(model_id, editing_rule_id)],
        }

    @staticmethod
    def create_model_unique_rule(model_id: str, data: dict[str, Any], username: str = "admin"):
        create_unique_rule(
            model_id,
            UniqueRulePayload(field_ids=list(data.get("field_ids") or [])),
            username,
        )
        return ModelManage.get_model_unique_rules(model_id)

    @staticmethod
    def update_model_unique_rule(
        model_id: str,
        rule_id: str,
        data: dict[str, Any],
        username: str = "admin",
    ):
        update_unique_rule(
            model_id,
            rule_id,
            UniqueRulePayload(field_ids=list(data.get("field_ids") or [])),
            username,
        )
        return ModelManage.get_model_unique_rules(model_id)

    @staticmethod
    def delete_model_unique_rule(model_id: str, rule_id: str, username: str = "admin"):
        delete_unique_rule(model_id, rule_id, username)
        return ModelManage.get_model_unique_rules(model_id)

    @staticmethod
    def model_association_create(**data):
        """
        创建模型关联
        """
        with GraphClient() as ag:
            try:
                edge = ag.create_edge(
                    MODEL_ASSOCIATION,
                    data["src_id"],
                    MODEL,
                    data["dst_id"],
                    MODEL,
                    data,
                    "model_asst_id",
                )
            except BaseAppException as e:
                if e.message == "edge already exists":
                    raise BaseAppException("model association repetition")
                else:
                    raise BaseAppException(e.message)
        return edge

    @staticmethod
    def model_association_delete(id: int):
        """
        删除模型关联
        """
        with GraphClient() as ag:
            ag.delete_edge(id)

    @staticmethod
    def model_association_info_search(model_asst_id: str):
        """
        查询模型关联详情
        """
        with GraphClient() as ag:
            query_data = {
                "field": "model_asst_id",
                "type": "str=",
                "value": model_asst_id,
            }
            edges = ag.query_edge(MODEL_ASSOCIATION, [query_data])
        if len(edges) == 0:
            return {}
        return edges[0]

    @staticmethod
    def model_association_search(model_id: str):
        """
        查询模型所有的关联
        """
        query_list = [
            {"field": "src_model_id", "type": "str=", "value": model_id},
            {"field": "dst_model_id", "type": "str=", "value": model_id},
        ]
        with GraphClient() as ag:
            edges = ag.query_edge(MODEL_ASSOCIATION, query_list, param_type="OR")

        return edges

    @staticmethod
    def check_model_exist_association(model_id):
        """模型存在关联关系"""
        edges = ModelManage.model_association_search(model_id)
        if edges:
            raise BaseAppException("model association exist")

    @staticmethod
    def check_model_exist_inst(model_id):
        """模型存在实例"""
        params = [{"field": "model_id", "type": "str=", "value": model_id}]
        with GraphClient() as ag:
            _, count = ag.query_entity(INSTANCE, params, page=dict(skip=0, limit=1))
        if count > 0:
            raise BaseAppException("model exist instance")

    # ===========

    @staticmethod
    def get_max_order_id(classification_id: str):
        """
        获取当前最大的 order_id
        Args:
            classification_id: 分类ID
        """
        with GraphClient() as ag:
            models, _ = ag.query_entity(
                MODEL,
                [
                    {
                        "field": "classification_id",
                        "type": "str=",
                        "value": classification_id,
                    }
                ],
                order="order_id",
                order_type="desc",
                page={"skip": 0, "limit": 1},
            )
            if not models:
                return 0
            return models[0].get("order_id", 0)

    @staticmethod
    def update_model_orders(model_orders: list):
        """
        批量更新模型排序
        Args:
            model_orders: [{"model_id": "model_1", "order_id": 1}, ...]
        """
        with GraphClient() as ag:
            for order_info in model_orders:
                model_query = {
                    "field": "model_id",
                    "type": "str=",
                    "value": order_info["model_id"],
                }
                models, model_count = ag.query_entity(MODEL, [model_query])
                if model_count == 0:
                    continue
                model_info = models[0]
                ag.set_entity_properties(
                    MODEL,
                    [model_info["_id"]],
                    {"order_id": order_info["order_id"]},
                    {},
                    [],
                    False,
                )
        return True

    @staticmethod
    def export_model_config(language):
        """导出模型配置为Excel (按model_config.xlsx模板格式)"""
        from io import BytesIO

        from openpyxl import Workbook

        CLASSIFICATION_HEADERS_CN = ["模型分类ID", "模型分类名称"]
        CLASSIFICATION_HEADERS_EN = ["classification_id", "classification_name"]

        MODEL_HEADERS_CN = ["模型ID", "模型名称", "模型图标", "模型分类ID"]
        MODEL_HEADERS_EN = ["model_id", "model_name", "icn", "classification_id"]

        ATTR_HEADERS_CN = [
            "英文名",
            "名称",
            "类型",
            "数据配置",
            "分组",
            "是否唯一",
            "是否可编辑",
            "是否必填",
            "用户提示",
            "唯一规则序号",
            "默认值",
        ]
        ATTR_HEADERS_EN = [
            "attr_id",
            "attr_name",
            "attr_type",
            "option",
            "attr_group",
            "is_only",
            "editable",
            "is_required",
            "user_prompt",
            "unique_rule_order",
            "default_value",
        ]

        ASSO_HEADERS_CN = ["源模型", "目标模型", "关联关系", "源-目标约束"]
        ASSO_HEADERS_EN = ["src_model_id", "dst_model_id", "asst_id", "mapping"]
        PUBLIC_ENUM_LIBRARY_HEADERS_CN = [
            "公共选项库ID",
            "公共选项库名称",
            "组织列表",
            "选项列表",
        ]
        PUBLIC_ENUM_LIBRARY_HEADERS_EN = [
            "library_id",
            "name",
            "team",
            "options",
        ]

        workbook = Workbook()

        ws_classifications = workbook.active
        if ws_classifications is None:
            ws_classifications = workbook.create_sheet(title="classifications")
        else:
            ws_classifications.title = "classifications"
        ws_classifications.append(CLASSIFICATION_HEADERS_CN)
        ws_classifications.append(CLASSIFICATION_HEADERS_EN)

        classifications = ClassificationManage.search_model_classification(language=language)
        for classification in classifications:
            ws_classifications.append(
                [
                    classification.get("classification_id", ""),
                    classification.get("classification_name", ""),
                ]
            )

        ws_models = workbook.create_sheet(title="models")
        ws_models.append(MODEL_HEADERS_CN)
        ws_models.append(MODEL_HEADERS_EN)

        ws_public_enum_libraries = workbook.create_sheet(title="public_enum_libraries")
        ws_public_enum_libraries.append(PUBLIC_ENUM_LIBRARY_HEADERS_CN)
        ws_public_enum_libraries.append(PUBLIC_ENUM_LIBRARY_HEADERS_EN)

        from apps.cmdb.services.public_enum_library import list_libraries

        for library in list_libraries():
            ws_public_enum_libraries.append(
                [
                    library.get("library_id", ""),
                    library.get("name", ""),
                    json.dumps(library.get("team", []), ensure_ascii=False),
                    json.dumps(library.get("options", []), ensure_ascii=False),
                ]
            )

        models = ModelManage.search_model(language=language)
        for model in models:
            ws_models.append(
                [
                    model.get("model_id", ""),
                    model.get("model_name", ""),
                    model.get("icn", ""),
                    model.get("classification_id", ""),
                ]
            )

        for model in models:
            model_id = model.get("model_id", "")
            if not model_id:
                continue

            ws_attr = workbook.create_sheet(title=f"attr-{model_id}")
            ws_attr.append(ATTR_HEADERS_CN)
            ws_attr.append(ATTR_HEADERS_EN)

            attrs = ModelManage._normalize_attr_constraints(ModelManage.parse_attrs(model.get("attrs", "[]")))
            unique_rules = build_unique_rule_context(model_id).unique_rules
            attr_rows = []
            for attr in attrs:
                if attr.get("is_display_field"):
                    continue
                option = attr.get("option", {})
                if attr.get("attr_type") == "enum":
                    if attr.get("enum_rule_type") == "public_library":
                        option = {
                            "enum_rule_type": "public_library",
                            "public_library_id": attr.get("public_library_id"),
                            "enum_select_mode": attr.get("enum_select_mode", ENUM_SELECT_MODE_DEFAULT),
                            "option": option if isinstance(option, list) else [],
                        }
                    elif isinstance(option, list):
                        option = option
                option_str = json.dumps(option, ensure_ascii=False) if option else ""
                attr_rows.append(
                    {
                        "attr_id": attr.get("attr_id", ""),
                        "attr_name": attr.get("attr_name", ""),
                        "attr_type": attr.get("attr_type", ""),
                        "option": option_str,
                        "attr_group": attr.get("attr_group", ""),
                        "is_only": attr.get("is_only", False),
                        "editable": attr.get("editable", True),
                        "is_required": attr.get("is_required", False),
                        "user_prompt": attr.get("user_prompt", ""),
                        "default_value": json.dumps(attr.get("default_value", []), ensure_ascii=False) if attr.get("attr_type") == "enum" else "",
                    }
                )

            attr_rows = apply_unique_rules_to_attr_export_rows(attr_rows, unique_rules)
            logger.info(
                "[UniqueRule] attr export complete model_id=%s sheet_name=%s rule_count=%s",
                model_id,
                f"attr-{model_id}",
                len(unique_rules),
            )
            for row in attr_rows:
                ws_attr.append(
                    [
                        row.get("attr_id", ""),
                        row.get("attr_name", ""),
                        row.get("attr_type", ""),
                        row.get("option", ""),
                        row.get("attr_group", ""),
                        row.get("is_only", False),
                        row.get("editable", True),
                        row.get("is_required", False),
                        row.get("user_prompt", ""),
                        row.get("unique_rule_order", ""),
                        row.get("default_value", ""),
                    ]
                )

            associations = ModelManage.model_association_search(model_id)
            if associations:
                ws_asso = workbook.create_sheet(title=f"asso-{model_id}")
                ws_asso.append(ASSO_HEADERS_CN)
                ws_asso.append(ASSO_HEADERS_EN)
                for asso in associations:
                    ws_asso.append(
                        [
                            asso.get("src_model_id", ""),
                            asso.get("dst_model_id", ""),
                            asso.get("asst_id", ""),
                            asso.get("mapping", ""),
                        ]
                    )

        file_stream = BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)

        return file_stream

    @staticmethod
    def import_model_config(file):
        from apps.cmdb.model_migrate.migrete_service import ModelMigrate

        migrator = ModelMigrate(file_source=file, is_pre=False)
        result = migrator.main()
        model_config = migrator.model_config
        with GraphClient() as ag:
            for sheet_name, rows in model_config.items():
                if not sheet_name.startswith("attr-"):
                    continue
                model_id = sheet_name.replace("attr-", "", 1)
                model_info = ModelManage.search_model_info(model_id)
                if not model_info:
                    continue
                try:
                    rules = build_unique_rules_from_attr_rows(model_id, rows)
                    exist_items, _ = ag.query_entity(
                        INSTANCE,
                        [{"field": "model_id", "type": "str=", "value": model_id}],
                    )
                    conflicts = validate_unique_rules_against_existing_instances(
                        model_id,
                        rules,
                        exist_items,
                    )
                    if conflicts:
                        raise BaseAppException(conflicts[0].message)
                    ag.set_entity_properties(
                        MODEL,
                        [model_info["_id"]],
                        {"unique_rules": json.dumps([asdict(rule) for rule in rules], ensure_ascii=False)},
                        {},
                        [],
                        False,
                    )
                    logger.info(
                        "[UniqueRule] attr import success model_id=%s sheet_name=%s rule_count=%s",
                        model_id,
                        sheet_name,
                        len(rules),
                    )
                except Exception as err:
                    logger.warning(
                        "[UniqueRule] attr import failed model_id=%s sheet_name=%s row_number=%s reason=%s",
                        model_id,
                        sheet_name,
                        0,
                        getattr(err, "message", str(err)),
                    )
                    raise
        return result
